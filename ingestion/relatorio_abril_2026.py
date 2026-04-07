"""
Gera relatorio_abril_2026.xlsx com 3 abas:
  1. resumo_diario   — conversas ativas, novos contatos e tempo de resposta por dia
  2. tempo_resposta  — detalhamento por conversa (mediana, p95, etc.)
  3. dados_brutos    — tabela completa de conversas com tempo de resposta calculado
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime, date
from statistics import median, mean, quantiles

BRT_FMT = "%Y-%m-%d %H:%M:%S"

# ── Carrega planilha fonte ────────────────────────────────────────────────────
SOURCE = "conversas_abril_2026.xlsx"
OUTPUT = "relatorio_abril_2026.xlsx"

wb_in = openpyxl.load_workbook(SOURCE)

ws_conv = wb_in["conversations"]
conv_headers = [c.value for c in ws_conv[1]]
convs = [dict(zip(conv_headers, row)) for row in ws_conv.iter_rows(min_row=2, values_only=True)]

ws_msg = wb_in["messages"]
msg_headers = [c.value for c in ws_msg[1]]
msgs = [dict(zip(msg_headers, row)) for row in ws_msg.iter_rows(min_row=2, values_only=True)]


# ── Calcula tempo de resposta por conversa ────────────────────────────────────
def parse_dt(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, BRT_FMT)
    except Exception:
        return None


msgs_by_conv = defaultdict(list)
for m in msgs:
    msgs_by_conv[m["conversation_id"]].append(m)


def first_response_seconds(conv_id):
    ms = sorted(msgs_by_conv.get(conv_id, []), key=lambda m: m["dateAdded_iso"] or "")
    first_in = None
    for m in ms:
        if m["direction"] == "inbound" and first_in is None:
            first_in = parse_dt(m["dateAdded_iso"])
        elif m["direction"] == "outbound" and first_in:
            t2 = parse_dt(m["dateAdded_iso"])
            if t2 and t2 >= first_in:
                return int((t2 - first_in).total_seconds())
            break
    return None


# Enriquece conversas com tempo de resposta
for c in convs:
    c["_resp_sec"] = first_response_seconds(c["id"])
    c["_day_activity"] = c["lastMessageDate_iso"][:10] if c["lastMessageDate_iso"] else None
    c["_day_added"] = c["dateAdded_iso"][:10] if c["dateAdded_iso"] else None


# ── Helpers de estilo ─────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1E40AF")   # azul escuro
ALT_FILL    = PatternFill("solid", fgColor="EFF6FF")   # azul muito claro
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=13, color="1E3A8A")

thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_row(ws, row_num, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_data_row(ws, row_num, ncols, alternate=False):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        if alternate:
            cell.fill = ALT_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)


def autofit(ws, min_w=10, max_w=45):
    for col in ws.columns:
        length = max((len(str(cell.value or "")) for cell in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


def fmt_duration(secs):
    if secs is None:
        return "—"
    if secs < 60:
        return f"{secs}s"
    if secs < 3600:
        return f"{secs // 60}min {secs % 60}s"
    h = secs // 3600
    m = (secs % 3600) // 60
    return f"{h}h {m}min"


# ── ABA 1: resumo_diario ──────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws1 = wb_out.active
ws1.title = "resumo_diario"
ws1.row_dimensions[1].height = 30

# Agrupa por dia
days_sorted = sorted({c["_day_activity"] for c in convs if c["_day_activity"]})

daily = {}
for d in days_sorted:
    convs_day = [c for c in convs if c["_day_activity"] == d]
    new_contacts = [c for c in convs if c["_day_added"] == d]

    resp_times = [c["_resp_sec"] for c in convs_day if c["_resp_sec"] is not None]
    # Filtra outliers acima de 24h para a média não distorcer
    resp_times_filtered = [r for r in resp_times if r <= 86400]

    daily[d] = {
        "conversas_ativas": len(convs_day),
        "novos_contatos": len(new_contacts),
        "com_resposta": len(resp_times),
        "mediana_resp": int(median(resp_times_filtered)) if resp_times_filtered else None,
        "media_resp": int(mean(resp_times_filtered)) if resp_times_filtered else None,
        "p95_resp": int(quantiles(resp_times_filtered, n=20)[18]) if len(resp_times_filtered) >= 5 else None,
    }

headers1 = [
    "Data",
    "Conversas ativas",
    "Novos contatos",
    "Com resposta",
    "Mediana resposta",
    "Média resposta",
    "P95 resposta",
]

ws1.append(headers1)
style_header_row(ws1, 1, len(headers1))

for i, d in enumerate(days_sorted, start=2):
    v = daily[d]
    row = [
        d,
        v["conversas_ativas"],
        v["novos_contatos"],
        v["com_resposta"],
        fmt_duration(v["mediana_resp"]),
        fmt_duration(v["media_resp"]),
        fmt_duration(v["p95_resp"]),
    ]
    ws1.append(row)
    style_data_row(ws1, i, len(headers1), alternate=(i % 2 == 0))

# Linha de totais
ws1.append([])
total_row = len(days_sorted) + 3
all_resp = [c["_resp_sec"] for c in convs if c["_resp_sec"] is not None and c["_resp_sec"] <= 86400]
totals = [
    "TOTAL / GERAL",
    sum(v["conversas_ativas"] for v in daily.values()),
    sum(v["novos_contatos"] for v in daily.values()),
    sum(v["com_resposta"] for v in daily.values()),
    fmt_duration(int(median(all_resp)) if all_resp else None),
    fmt_duration(int(mean(all_resp)) if all_resp else None),
    fmt_duration(int(quantiles(all_resp, n=20)[18]) if len(all_resp) >= 5 else None),
]
ws1.append(totals)
for col in range(1, len(headers1) + 1):
    cell = ws1.cell(row=total_row, column=col)
    cell.font = Font(bold=True, color="1E3A8A")
    cell.fill = PatternFill("solid", fgColor="DBEAFE")
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center")

autofit(ws1)


# ── ABA 2: tempo_resposta ─────────────────────────────────────────────────────
ws2 = wb_out.create_sheet("tempo_resposta")
ws2.row_dimensions[1].height = 30

# Detalhe por conversa com tempo de resposta
headers2 = [
    "Data atividade",
    "Contato",
    "Telefone",
    "Canal",
    "Status",
    "Tags",
    "Mensagens totais",
    "Tempo 1ª resposta",
    "Tempo (segundos)",
]

ws2.append(headers2)
style_header_row(ws2, 1, len(headers2))

msg_count_by_conv = {conv_id: len(ms) for conv_id, ms in msgs_by_conv.items()}

rows2 = sorted(
    [c for c in convs if c["_resp_sec"] is not None],
    key=lambda c: c["_day_activity"] or ""
)

for i, c in enumerate(rows2, start=2):
    row = [
        c["_day_activity"],
        c["contactName"] or c["fullName"] or "",
        c["phone"] or "",
        c["type"] or "",
        c["lastMessageType"] or "",
        c["tags"] or "",
        msg_count_by_conv.get(c["id"], 0),
        fmt_duration(c["_resp_sec"]),
        c["_resp_sec"],
    ]
    ws2.append(row)
    style_data_row(ws2, i, len(headers2), alternate=(i % 2 == 0))

autofit(ws2)


# ── ABA 3: dados_brutos ───────────────────────────────────────────────────────
ws3 = wb_out.create_sheet("dados_brutos")
ws3.row_dimensions[1].height = 30

headers3 = [
    "id", "Data atividade", "Data criacao", "Contato", "Telefone",
    "Email", "Canal", "Ultimo tipo msg", "Ultimo corpo msg",
    "Tags", "Atribuido a", "Mensagens", "Tempo 1a resposta", "Tempo (seg)",
]

ws3.append(headers3)
style_header_row(ws3, 1, len(headers3))

for i, c in enumerate(sorted(convs, key=lambda c: c["_day_activity"] or ""), start=2):
    row = [
        c["id"],
        c["_day_activity"],
        c["_day_added"],
        c["contactName"] or c["fullName"] or "",
        c["phone"] or "",
        c["email"] or "",
        c["type"] or "",
        c["lastMessageType"] or "",
        (c["lastMessageBody"] or "")[:120],
        c["tags"] or "",
        c["assigned_to"] or "",
        msg_count_by_conv.get(c["id"], 0),
        fmt_duration(c["_resp_sec"]),
        c["_resp_sec"],
    ]
    ws3.append(row)
    style_data_row(ws3, i, len(headers3), alternate=(i % 2 == 0))

autofit(ws3)

wb_out.save(OUTPUT)
print(f"[relatorio] Salvo em: {OUTPUT}")
print(f"  Dias cobertos: {len(days_sorted)}")
print(f"  Total conversas: {len(convs)}")
print(f"  Com tempo de resposta: {len([c for c in convs if c['_resp_sec'] is not None])}")
for d in days_sorted:
    v = daily[d]
    print(f"  {d} — ativas: {v['conversas_ativas']} | novos: {v['novos_contatos']} | mediana resp: {fmt_duration(v['mediana_resp'])}")
