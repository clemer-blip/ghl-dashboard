"""
Exporta todas as conversas com atividade em abril/2026 (horário de Brasília, UTC-3)
para um arquivo Excel com duas abas:
  - conversations: dados completos de cada conversa
  - messages: todas as mensagens dessas conversas
"""

import os
import sys
from datetime import datetime, timezone, timedelta
from dotenv import load_dotenv

load_dotenv()

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

from ghl_client import GHLClient

# ── Janela de abril/2026 em BRT (UTC-3) ──────────────────────────────────────
BRT = timezone(timedelta(hours=-3))
MONTH_START = datetime(2026, 4, 1, 0, 0, 0, tzinfo=BRT)
MONTH_END   = datetime(2026, 4, 30, 23, 59, 59, tzinfo=BRT)

# startAfterDate precisa ser um instante ANTES da janela
CURSOR_START = MONTH_START - timedelta(seconds=1)

# ── Campos que queremos das conversas ────────────────────────────────────────
CONV_FIELDS = [
    "id", "contactId", "locationId", "contactName", "fullName",
    "email", "phone", "type", "lastMessageBody", "lastMessageType",
    "lastMessageDate_iso", "unreadCount", "assigned_to", "inbox",
    "starred", "tags", "source", "attributionSource",
    "dateAdded_iso", "dateUpdated_iso",
]

MSG_FIELDS = [
    "conversation_id", "contact_name", "id", "type", "messageType",
    "direction", "status", "body", "dateAdded_iso",
    "attachments", "userId", "contactId",
]

DIRECTION_MAP = {1: "inbound", 2: "outbound"}


def ts_to_iso(ts_ms):
    if not ts_ms:
        return ""
    if isinstance(ts_ms, str):
        try:
            dt = datetime.fromisoformat(ts_ms.replace("Z", "+00:00"))
            return dt.astimezone(BRT).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return ts_ms
    return datetime.fromtimestamp(ts_ms / 1000, tz=BRT).strftime("%Y-%m-%d %H:%M:%S")


def in_window(ts_ms):
    if not ts_ms:
        return False
    dt = datetime.fromtimestamp(ts_ms / 1000, tz=BRT)
    return MONTH_START <= dt <= MONTH_END


def flatten_conv(conv):
    return {
        "id":                  conv.get("id", ""),
        "contactId":           conv.get("contactId", ""),
        "locationId":          conv.get("locationId", ""),
        "contactName":         conv.get("contactName", ""),
        "fullName":            conv.get("fullName", ""),
        "email":               conv.get("email", ""),
        "phone":               conv.get("phone", ""),
        "type":                conv.get("type", ""),
        "lastMessageBody":     conv.get("lastMessageBody", ""),
        "lastMessageType":     conv.get("lastMessageType", ""),
        "lastMessageDate_iso": ts_to_iso(conv.get("lastMessageDate")),
        "unreadCount":         conv.get("unreadCount", 0),
        "assigned_to":         conv.get("assignedTo", conv.get("userId", "")),
        "inbox":               conv.get("inbox", ""),
        "starred":             conv.get("starred", False),
        "tags":                ", ".join(conv.get("tags", []) or []),
        "source":              conv.get("source", ""),
        "attributionSource":   str(conv.get("attributionSource", "") or ""),
        "dateAdded_iso":       ts_to_iso(conv.get("dateAdded")),
        "dateUpdated_iso":     ts_to_iso(conv.get("dateUpdated")),
    }


def flatten_msg(msg, conv_id, contact_name):
    direction_raw = msg.get("direction", "")
    if isinstance(direction_raw, int):
        direction = DIRECTION_MAP.get(direction_raw, str(direction_raw))
    else:
        direction = str(direction_raw)

    attachments = msg.get("attachments", []) or []
    att_str = ", ".join(
        a.get("url", str(a)) if isinstance(a, dict) else str(a)
        for a in attachments
    )

    return {
        "conversation_id": conv_id,
        "contact_name":    contact_name,
        "id":              msg.get("id", ""),
        "type":            msg.get("type", ""),
        "messageType":     msg.get("messageType", ""),
        "direction":       direction,
        "status":          msg.get("status", ""),
        "body":            msg.get("body", ""),
        "dateAdded_iso":   ts_to_iso(msg.get("dateAdded")),
        "attachments":     att_str,
        "userId":          msg.get("userId", ""),
        "contactId":       msg.get("contactId", ""),
    }


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def main():
    ghl = GHLClient(
        api_key=os.environ["GHL_API_KEY"],
        location_id=os.environ["GHL_LOCATION_ID"],
    )

    print(f"[export] Buscando conversas de abril/2026 (BRT)...")

    conversations = []
    all_messages  = []
    skipped_after = 0

    for conv in ghl.iter_conversations(start_after_date=CURSOR_START):
        ts = conv.get("lastMessageDate", 0)
        dt = datetime.fromtimestamp(ts / 1000, tz=BRT) if ts else None

        # Conversas além do mês → podemos parar
        if dt and dt > MONTH_END:
            skipped_after += 1
            if skipped_after >= 3:
                break
            continue

        if not in_window(ts):
            continue

        contact_name = conv.get("contactName") or conv.get("fullName") or ""
        print(f"  → {contact_name or conv['id']} | {ts_to_iso(ts)}")

        conversations.append(flatten_conv(conv))

        messages = ghl.get_messages(conv["id"])
        for msg in messages:
            all_messages.append(flatten_msg(msg, conv["id"], contact_name))

    print(f"[export] {len(conversations)} conversas | {len(all_messages)} mensagens")

    # ── Excel ─────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    ws_conv = wb.active
    ws_conv.title = "conversations"
    write_sheet(ws_conv, CONV_FIELDS, conversations)

    ws_msg = wb.create_sheet("messages")
    write_sheet(ws_msg, MSG_FIELDS, all_messages)

    output_path = os.path.join(
        os.path.dirname(__file__), "conversas_abril_2026.xlsx"
    )
    wb.save(output_path)
    print(f"[export] Salvo em: {output_path}")


if __name__ == "__main__":
    main()
